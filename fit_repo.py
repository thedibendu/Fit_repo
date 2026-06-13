#[Publichow to run
# from fit_repo import run_hill_fit_multiple_from_excel as mfit
# result = mfit("/Users/dibendu/Documents/atp_analysis/Book2.xlsx")

# to run the array form
# from fit_repo import run_hill_fit_from_arrays as afit
# result = afit(substrate_conc, initial_rates)

# to run the automated raw time/absorbance form
# from fit_repo import run_hill_fit_from_arrays as afit
# result = afit(
#     substrate_conc,
#     time=time,
#     absorbance_data=absorbance_data,
#     fit_start=0,
#     fit_end=60
# )

import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import openpyxl
import argparse
import os
from scipy.stats import linregress


def load_excel_data(filepath, x_col=1, y_col=2, sheet_name=None, skip_header=True):
    """
    Load x and y data from an Excel sheet for a single dataset.

    Parameters
    ----------
    filepath : str
        Path to the Excel file.
    x_col : int
        Zero-based column index for x data.
    y_col : int
        Zero-based column index for y data.
    sheet_name : str or None
        Sheet name to use. If None, uses active sheet.
    skip_header : bool
        Whether to skip the first row.

    Returns
    -------
    x : np.ndarray
    y : np.ndarray
    """

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    x_data = []
    y_data = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if skip_header and i == 0:
            continue

        if row[x_col] is not None and row[y_col] is not None:
            x_data.append(row[x_col])
            y_data.append(row[y_col])

    return np.array(x_data, dtype=float), np.array(y_data, dtype=float)


def load_excel_grouped_data(
    filepath,
    condition_col=0,
    x_col=1,
    y_col=2,
    sheet_name=None,
    skip_header=True
):
    """
    Load x and y data from an Excel sheet and group by condition.

    Expected Excel format
    ---------------------
    Condition | ATP | Rate

    Example
    -------
    A | 0.1 | 0.5
    A | 0.2 | 0.9
    B | 0.1 | 0.3
    B | 0.2 | 0.6

    Parameters
    ----------
    filepath : str
        Path to Excel file.
    condition_col : int
        Zero-based column index for condition labels.
    x_col : int
        Zero-based column index for x data.
    y_col : int
        Zero-based column index for y data.
    sheet_name : str or None
        Sheet name to use. If None, uses active sheet.
    skip_header : bool
        Whether to skip the first row.

    Returns
    -------
    grouped_data : dict
        Dictionary like:
        {
            "A": {"x": np.ndarray, "y": np.ndarray},
            "B": {"x": np.ndarray, "y": np.ndarray}
        }
    """

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    grouped_data = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if skip_header and i == 0:
            continue

        if (
            row[condition_col] is None
            or row[x_col] is None
            or row[y_col] is None
        ):
            continue

        condition = str(row[condition_col]).strip()

        if condition not in grouped_data:
            grouped_data[condition] = {"x": [], "y": []}

        grouped_data[condition]["x"].append(row[x_col])
        grouped_data[condition]["y"].append(row[y_col])

    for condition in grouped_data:
        grouped_data[condition]["x"] = np.array(
            grouped_data[condition]["x"],
            dtype=float
        )
        grouped_data[condition]["y"] = np.array(
            grouped_data[condition]["y"],
            dtype=float
        )

    return grouped_data


def hill_equation(S, Vmax, Khalf, n):
    """
    Hill equation.
    """

    return (Vmax * (S ** n)) / ((Khalf ** n) + (S ** n))


def fit_hill_equation(substrate_conc, initial_rates):
    """
    Fit Hill equation to data.

    Returns
    -------
    popt : tuple
        Best-fit parameters: Vmax, Khalf, n
    perr : np.ndarray
        Standard errors of fitted parameters
    """

    p0 = [
        max(initial_rates),
        np.median(substrate_conc),
        1.0
    ]

    bounds = (0, np.inf)

    popt, pcov = curve_fit(
        hill_equation,
        substrate_conc,
        initial_rates,
        p0=p0,
        bounds=bounds,
        maxfev=20000
    )

    perr = np.sqrt(np.diag(pcov))

    return popt, perr


def calculate_model_and_errors(substrate_conc, initial_rates, Vmax, Khalf, n):
    """
    Calculate model prediction, residuals, and absolute residual-based errors.
    """

    y_model = hill_equation(substrate_conc, Vmax, Khalf, n)
    residuals = initial_rates - y_model
    errors = np.abs(residuals)

    return y_model, residuals, errors


def generate_fit_curve(substrate_conc, Vmax, Khalf, n, num_points=500, x_scale=1.3):
    """
    Generate smooth fitted curve for plotting.
    """

    x_fit = np.linspace(0, max(substrate_conc) * x_scale, num_points)
    y_fit = hill_equation(x_fit, Vmax, Khalf, n)

    return x_fit, y_fit


def plot_hill_fit(
    substrate_conc,
    initial_rates,
    x_fit,
    y_fit,
    errors=None,
    output_file=None,
    point_style="o",
    curve_style=":",
    color=None
):
    """
    Plot one dataset and one fitted Hill curve.
    Data points and fit curve use the same color.
    """

    plt.figure(figsize=(8, 6))

    if errors is not None:
        plt.errorbar(
            substrate_conc,
            initial_rates,
            yerr=errors,
            fmt=point_style,
            capsize=5,
            color=color
        )
    else:
        plt.plot(
            substrate_conc,
            initial_rates,
            point_style,
            color=color
        )

    plt.plot(
        x_fit,
        y_fit,
        curve_style,
        linewidth=1.54,
        color=color
    )

    plt.xlabel("[ATP] / mM", fontsize=18)
    plt.ylabel("Initial Rate (µM/s)", fontsize=18)
    plt.tick_params(axis="both", labelsize=18)
    plt.tight_layout()
    if output_file:
        plt.savefig(output_file, dpi=300)
    plt.show()
    plt.close()


def plot_multiple_hill_fits(
    fit_results_dict,
    output_file=None
):
    """
    Plot multiple condition datasets and their fitted curves on the same graph.
    Each condition uses one matching color for both data points and fitted curve.
    """

    plt.figure(figsize=(8, 6))

    markers = ["o", "o", "o", "D", "v", "P", "X", "*"]
    colors = ["red", "darkgreen", "blue", "black"]

    for i, (condition, result) in enumerate(fit_results_dict.items()):
        marker = markers[i % len(markers)]
        color = colors[i % len(colors)]

        plt.errorbar(
            result["substrate_conc"],
            result["initial_rates"],
            yerr=result["errors"],
            fmt=marker,
            capsize=5,
            color=color,
            label=f"{condition}"
        )

        plt.plot(
            result["x_fit"],
            result["y_fit"],
            linestyle=":",
            linewidth=1.54,
            color=color
        )

    plt.xlabel("[ATP] / mM", fontsize=18)
    plt.ylabel("Initial Rate (µM/s)", fontsize=18)
    plt.tick_params(axis="both", labelsize=18)
    plt.legend(fontsize=12,frameon= False)
    plt.tight_layout()
    if output_file:
        plt.savefig(output_file, dpi=300)
    plt.show()
    plt.close()


def print_fit_results(Vmax, Khalf, n, perr, condition=None):
    """
    Print fitted parameters neatly.
    """

    if condition is not None:
        print(f"\nFit results for condition: {condition}")
    else:
        print("\nFit results:")

    print(f"Vmax   = {Vmax:.6f} ± {perr[0]:.6f}")
    print(f"Khalf  = {Khalf:.6f} ± {perr[1]:.6f}")
    print(f"Hill n = {n:.6f} ± {perr[2]:.6f}")


def run_hill_fit_from_excel(
    filepath,
    x_col=1,
    y_col=2,
    sheet_name=None,
    output_file=None,
    color=None
):
    """
    Full pipeline for one dataset:
    load data -> fit Hill equation -> calculate errors -> plot -> print results
    """

    substrate_conc, initial_rates = load_excel_data(
        filepath,
        x_col=x_col,
        y_col=y_col,
        sheet_name=sheet_name
    )

    popt, perr = fit_hill_equation(substrate_conc, initial_rates)
    Vmax, Khalf, n = popt

    _, residuals, errors = calculate_model_and_errors(
        substrate_conc,
        initial_rates,
        Vmax,
        Khalf,
        n
    )

    x_fit, y_fit = generate_fit_curve(substrate_conc, Vmax, Khalf, n)

    print_fit_results(Vmax, Khalf, n, perr)

    plot_hill_fit(
        substrate_conc,
        initial_rates,
        x_fit,
        y_fit,
        errors=errors,
        output_file=output_file,
        color=color
    )

    return {
        "substrate_conc": substrate_conc,
        "initial_rates": initial_rates,
        "Vmax": Vmax,
        "Khalf": Khalf,
        "n": n,
        "parameter_errors": perr,
        "residuals": residuals
    }


# ======================================================================
# THIS IS THE ONLY UPDATED SECTION
# ======================================================================
def run_hill_fit_from_arrays(
    substrate_conc,
    initial_rates=None,
    time=None,
    absorbance_data=None,
    fit_start=None,
    fit_end=None,
    max_points=None,
    calibration_slope=0.0056,
    calibration_intercept=-0.0094,
    output_file=None,
    color=None
):
    """
    Full pipeline for ATPase Hill fitting.

    This function works in three modes.

    ------------------------------------------------------------
    MODE 1: OLD ARRAY WAY
    ------------------------------------------------------------
    Use this when you already have ATP concentrations and initial rates.

    Example:
    substrate_conc = [0.2, 0.4, 0.8, 1.6]
    initial_rates = [0.01, 0.03, 0.05, 0.08]

    result = run_hill_fit_from_arrays(substrate_conc, initial_rates)

    In this mode, initial_rates are assumed to already be in µM/s.

    ------------------------------------------------------------
    MODE 2: RAW ARRAY WAY
    ------------------------------------------------------------
    Use this when you have raw time and absorbance data.

    Example:
    substrate_conc = [0.2, 0.4, 0.8, 1.6]

    time = [0, 10, 20, 30, 40]

    absorbance_data = [
        [0.10, 0.11, 0.12, 0.13],
        [0.12, 0.15, 0.18, 0.25],
        [0.14, 0.19, 0.25, 0.37],
        [0.16, 0.23, 0.32, 0.49],
        [0.18, 0.27, 0.39, 0.61],
    ]

    result = run_hill_fit_from_arrays(
        substrate_conc,
        time=time,
        absorbance_data=absorbance_data,
        fit_start=0,
        fit_end=30
    )

    absorbance_data should be arranged as:

    rows    = time points
    columns = ATP concentrations

    ------------------------------------------------------------
    MODE 3: RAW EXCEL WAY
    ------------------------------------------------------------
    Use this when your Excel file is arranged like this:

    Time | 0.2 | 0.4 | 0.8 | 1.6 | 3.2 | 6.4
    0    | ...
    10   | ...
    20   | ...

    Example:
    result = run_hill_fit_from_arrays(
        "/Users/dibendu/Documents/atp_analysis/b4.xlsx",
        max_points=3
    )

    ------------------------------------------------------------
    Calibration
    ------------------------------------------------------------
    The raw ATPase slope is:

        slope_OD_per_s = OD/s

    The conversion used here is exactly:

        rate_uM_per_s = slope_OD_per_s / 0.0056

    The Hill fit uses rate_uM_per_s.
    """

    # ------------------------------------------------------------
    # MODE 3:
    # If substrate_conc is actually an Excel file path,
    # read ATP concentrations, time, and absorbance data from Excel.
    # ------------------------------------------------------------

    if isinstance(substrate_conc, str):

        filepath = substrate_conc

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError("Excel file must contain a header row and data rows.")

        header = rows[0]

        # First column is Time.
        # Every other column header is ATP concentration.
        substrate_conc = []

        for value in header[1:]:

            if value is None:
                continue

            if isinstance(value, str):
                cleaned = (
                    value.replace("mM", "")
                    .replace("ATP", "")
                    .replace("[", "")
                    .replace("]", "")
                    .strip()
                )
                substrate_conc.append(float(cleaned))
            else:
                substrate_conc.append(float(value))

        time = []
        absorbance_rows = []

        for row in rows[1:]:

            if row[0] is None:
                continue

            current_time = row[0]
            current_abs = row[1:1 + len(substrate_conc)]

            if all(v is None for v in current_abs):
                continue

            time.append(current_time)
            absorbance_rows.append(current_abs)

        substrate_conc = np.array(substrate_conc, dtype=float)
        time = np.array(time, dtype=float)
        absorbance_data = np.array(absorbance_rows, dtype=float)

        initial_rates = None

    else:
        substrate_conc = np.array(substrate_conc, dtype=float)

    # ------------------------------------------------------------
    # MODE 2 or MODE 3:
    # Calculate initial rates from raw OD/time data.
    # ------------------------------------------------------------

    if initial_rates is None:

        if time is None or absorbance_data is None:
            raise ValueError(
                "You must provide either initial_rates OR both time and absorbance_data."
            )

        time = np.array(time, dtype=float)
        absorbance_data = np.array(absorbance_data, dtype=float)

        if absorbance_data.ndim == 1:
            absorbance_data = absorbance_data.reshape(-1, 1)

        if absorbance_data.shape[0] != len(time):
            raise ValueError(
                "Number of rows in absorbance_data must match number of time points."
            )

        if absorbance_data.shape[1] != len(substrate_conc):
            raise ValueError(
                "Number of absorbance columns must match number of ATP concentrations."
            )

        calculated_rates = []
        rate_details = []

        print("\nCalculated initial rates:")
        print(
            "ATP_mM\t"
            "Slope_OD/s\t"
            "Rate_uM/s\t"
            "R_squared"
        )

        for i, atp in enumerate(substrate_conc):

            y = absorbance_data[:, i]

            valid = np.isfinite(time) & np.isfinite(y)

            x_valid = time[valid]
            y_valid = y[valid]

            if len(x_valid) < 2:
                raise ValueError(
                    f"Not enough valid points for ATP = {atp}"
                )

            # Sort by time
            order = np.argsort(x_valid)
            x_valid = x_valid[order]
            y_valid = y_valid[order]

            # Select initial linear region
            mask = np.ones_like(x_valid, dtype=bool)

            if fit_start is not None:
                mask = mask & (x_valid >= fit_start)

            if fit_end is not None:
                mask = mask & (x_valid <= fit_end)

            x_fit_region = x_valid[mask]
            y_fit_region = y_valid[mask]

            if max_points is not None:
                x_fit_region = x_fit_region[:max_points]
                y_fit_region = y_fit_region[:max_points]

            if len(x_fit_region) < 2:
                raise ValueError(
                    f"Not enough points to calculate initial rate for ATP = {atp}"
                )

            # Linear fit:
            # OD = slope * time + intercept
            slope_OD_per_s, intercept = np.polyfit(
                x_fit_region,
                y_fit_region,
                1
            )

            y_pred = slope_OD_per_s * x_fit_region + intercept

            ss_res = np.sum((y_fit_region - y_pred) ** 2)
            ss_tot = np.sum((y_fit_region - np.mean(y_fit_region)) ** 2)

            if ss_tot == 0:
                r_squared = np.nan
            else:
                r_squared = 1 - (ss_res / ss_tot)

            # ----------------------------------------------------
            # Final conversion requested:
            #
            # slope_OD_per_s = OD/s
            #
            # rate_uM_per_s = slope_OD_per_s / 0.0056
            #
            # This rate is directly used for Hill fitting.
            # ----------------------------------------------------

            rate_uM_per_s = slope_OD_per_s / calibration_slope

            # This is the rate used for Hill fitting.
            rate = rate_uM_per_s

            calculated_rates.append(rate)

            rate_details.append(
                {
                    "substrate_conc": atp,
                    "slope_OD_per_s": slope_OD_per_s,
                    "calibration_slope": calibration_slope,
                    "calibration_intercept": calibration_intercept,
                    "rate_uM_per_s": rate_uM_per_s,
                    "initial_rate": rate_uM_per_s,
                    "intercept": intercept,
                    "r_squared": r_squared,
                    "time_used": x_fit_region,
                    "absorbance_used": y_fit_region
                }
            )

            print(
                f"{atp:.6f}\t"
                f"{slope_OD_per_s:.6f}\t"
                f"{rate_uM_per_s:.9f}\t"
                f"{r_squared:.4f}"
            )

        initial_rates = np.array(calculated_rates, dtype=float)

    # ------------------------------------------------------------
    # MODE 1:
    # Use provided initial rates directly.
    # ------------------------------------------------------------

    else:
        initial_rates = np.array(initial_rates, dtype=float)
        rate_details = None

    # ------------------------------------------------------------
    # Original Hill fitting part.
    # Hill fit uses initial_rates in µM/s.
    # ------------------------------------------------------------

    popt, perr = fit_hill_equation(substrate_conc, initial_rates)
    Vmax, Khalf, n = popt

    _, residuals, errors = calculate_model_and_errors(
        substrate_conc,
        initial_rates,
        Vmax,
        Khalf,
        n
    )

    x_fit, y_fit = generate_fit_curve(
        substrate_conc,
        Vmax,
        Khalf,
        n
    )

    print_fit_results(Vmax, Khalf, n, perr)

    plot_hill_fit(
        substrate_conc,
        initial_rates,
        x_fit,
        y_fit,
        errors=errors,
        output_file=output_file,
        color=color
    )

    return {
        "substrate_conc": substrate_conc,
        "initial_rates": initial_rates,
        "Vmax": Vmax,
        "Khalf": Khalf,
        "n": n,
        "parameter_errors": perr,
        "residuals": residuals,
        "errors": errors,
        "x_fit": x_fit,
        "y_fit": y_fit,
        "rate_details": rate_details
    }
# ======================================================================
# EVERYTHING BELOW THIS PART IS SAME LOGIC AS BEFORE
# ======================================================================

def run_hill_fit_multiple_from_excel(
    filepath,
    condition_col=0,
    x_col=1,
    y_col=2,
    sheet_name=None,
    output_file=None
):
    """
    Full pipeline for grouped Excel data:
    load grouped data -> fit each condition separately -> plot together -> print results

    Expected Excel format:
    Condition | ATP | Rate
    """

    grouped_data = load_excel_grouped_data(
        filepath,
        condition_col=condition_col,
        x_col=x_col,
        y_col=y_col,
        sheet_name=sheet_name
    )

    all_results = {}

    for condition, data in grouped_data.items():
        substrate_conc = data["x"]
        initial_rates = data["y"]

        popt, perr = fit_hill_equation(substrate_conc, initial_rates)
        Vmax, Khalf, n = popt

        _, residuals, errors = calculate_model_and_errors(
            substrate_conc,
            initial_rates,
            Vmax,
            Khalf,
            n
        )

        x_fit, y_fit = generate_fit_curve(substrate_conc, Vmax, Khalf, n)

        print_fit_results(Vmax, Khalf, n, perr, condition=condition)

        all_results[condition] = {
            "substrate_conc": substrate_conc,
            "initial_rates": initial_rates,
            "Vmax": Vmax,
            "Khalf": Khalf,
            "n": n,
            "parameter_errors": perr,
            "residuals": residuals,
            "errors": errors,
            "x_fit": x_fit,
            "y_fit": y_fit
        }

    plot_multiple_hill_fits(
        all_results,
        output_file=output_file
    )

    return all_results


def load_raw_timecourse_data(filepath, sheet_name=None):
    """
    Load raw time-course data from Excel.
    Expects first column to be Time and subsequent columns to be OD values
    for different ATP concentrations.
    """

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Excel file must contain a header row and data rows.")

    header = rows[0]
    data_rows = rows[1:]

    time = []
    raw_data = {}
    col_to_atp = {}

    for col_idx, value in enumerate(header[1:], start=1):

        if value is None:
            continue

        try:
            atp_val_str = "".join(
                c for c in str(value) if c.isdigit() or c == "."
            )
            atp_conc = float(atp_val_str)
            col_to_atp[col_idx] = atp_conc
            raw_data[atp_conc] = []
        except ValueError:
            print(f"Warning: Could not parse ATP concentration from header: {value}")

    for row in data_rows:

        if row[0] is None:
            continue

        time.append(float(row[0]))

        for col_idx, atp_conc in col_to_atp.items():
            val = row[col_idx] if row[col_idx] is not None else 0.0
            raw_data[atp_conc].append(float(val))

    return np.array(time), raw_data


def select_initial_rate_region(time, od_values, fit_start=None, fit_end=None, max_points=3):
    """
    Select the time and OD points used for the initial-rate linear fit.
    """

    x_valid = np.array(time, dtype=float)
    y_valid = np.array(od_values, dtype=float)

    valid = np.isfinite(x_valid) & np.isfinite(y_valid)
    x_valid = x_valid[valid]
    y_valid = y_valid[valid]

    order = np.argsort(x_valid)
    x_valid = x_valid[order]
    y_valid = y_valid[order]

    mask = np.ones_like(x_valid, dtype=bool)

    if fit_start is not None:
        mask = mask & (x_valid >= fit_start)

    if fit_end is not None:
        mask = mask & (x_valid <= fit_end)

    x_fit_region = x_valid[mask]
    y_fit_region = y_valid[mask]

    if max_points is not None:
        x_fit_region = x_fit_region[:max_points]
        y_fit_region = y_fit_region[:max_points]

    return x_fit_region, y_fit_region


def calculate_initial_rates(
    time,
    raw_data,
    calibration_factor=0.0056,
    fit_start=None,
    fit_end=None,
    max_points=3
):
    """
    Calculate initial rates using linear regression for each ATP concentration.
    Returns sorted arrays of substrate concentrations and initial rates.
    """

    substrate_concs = []
    initial_rates = []

    print("\nCalculated Initial Rates from Raw Data:")
    print("-" * 50)
    print(f"{'ATP (mM)':<10} | {'Slope (OD/s)':<15} | {'Rate (uM/s)':<12} | {'R^2':<8}")
    print("-" * 50)

    for atp, od_values in raw_data.items():
        x_fit_region, y_fit_region = select_initial_rate_region(
            time,
            od_values,
            fit_start=fit_start,
            fit_end=fit_end,
            max_points=max_points
        )

        if len(x_fit_region) < 2:
            raise ValueError(
                f"Not enough points to calculate initial rate for ATP = {atp}"
            )

        slope, intercept, r_value, p_value, std_err = linregress(
            x_fit_region,
            y_fit_region
        )
        rate_uM_s = slope / calibration_factor

        substrate_concs.append(atp)
        initial_rates.append(rate_uM_s)

        print(
            f"{atp:<10.3f} | "
            f"{slope:<15.6f} | "
            f"{rate_uM_s:<12.4f} | "
            f"{r_value ** 2:<8.4f}"
        )

    substrate_concs = np.array(substrate_concs)
    initial_rates = np.array(initial_rates)
    idx = np.argsort(substrate_concs)

    return substrate_concs[idx], initial_rates[idx]


def run_hill_fit_fully_automated(
    filepath,
    sheet_name=None,
    calibration_factor=0.0056,
    fit_start=None,
    fit_end=None,
    max_points=3,
    output_prefix=None
):
    """
    Complete automated pipeline:
    raw Excel -> initial rates -> Hill fit -> raw time-course plot -> Hill plot.
    """

    time, raw_data = load_raw_timecourse_data(filepath, sheet_name)
    substrate_conc, initial_rates = calculate_initial_rates(
        time,
        raw_data,
        calibration_factor,
        fit_start=fit_start,
        fit_end=fit_end,
        max_points=max_points
    )

    popt, perr = fit_hill_equation(substrate_conc, initial_rates)
    Vmax, Khalf, n = popt

    _, residuals, errors = calculate_model_and_errors(
        substrate_conc,
        initial_rates,
        Vmax,
        Khalf,
        n
    )

    x_fit, y_fit = generate_fit_curve(substrate_conc, Vmax, Khalf, n)

    print_fit_results(Vmax, Khalf, n, perr)

    script_dir = os.path.dirname(os.path.abspath(filepath))

    plt.figure(figsize=(10, 6))

    for atp, od_values in raw_data.items():
        x_fit_region, y_fit_region = select_initial_rate_region(
            time,
            od_values,
            fit_start=fit_start,
            fit_end=fit_end,
            max_points=max_points
        )

        slope, intercept, _, _, _ = linregress(x_fit_region, y_fit_region)

        plt.scatter(time, od_values, alpha=0.35, label=f"{atp} mM")
        plt.scatter(x_fit_region, y_fit_region, edgecolors="black", linewidths=0.7)
        plt.plot(
            x_fit_region,
            intercept + slope * x_fit_region,
            "--",
            alpha=0.8
        )

    plt.xlabel("Time (s)")
    plt.ylabel("Absorbance (OD)")
    plt.title("Validation: Raw Time Course Linear Fits")
    plt.legend(bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.tight_layout()

    if output_prefix:
        plt.savefig(
            os.path.join(script_dir, f"{output_prefix}time_courses.png"),
            dpi=300
        )

    plt.show()
    plt.close()

    hill_fit_output_file = None

    if output_prefix:
        hill_fit_output_file = os.path.join(
            script_dir,
            f"{output_prefix}hill_fit.png"
        )

    plot_hill_fit(
        substrate_conc,
        initial_rates,
        x_fit,
        y_fit,
        errors=errors,
        output_file=hill_fit_output_file,
        color="blue"
    )

    if output_prefix:
        print(f"\nPlots saved to: {script_dir}")
        print(f"1. {output_prefix}time_courses.png")
        print(f"2. {output_prefix}hill_fit.png")

    return {
        "substrate_conc": substrate_conc,
        "initial_rates": initial_rates,
        "Vmax": Vmax,
        "Khalf": Khalf,
        "n": n,
        "parameter_errors": perr,
        "residuals": residuals,
        "errors": errors,
        "x_fit": x_fit,
        "y_fit": y_fit
    }


def build_cli_parser():
    """
    Build the command-line parser for running fits without editing this file.
    """

    parser = argparse.ArgumentParser(
        description="Run ATPase Hill equation fits from Excel files."
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    raw_parser = subparsers.add_parser(
        "raw",
        help="Fit from raw Excel time-course data: Time | ATP1 | ATP2 | ..."
    )
    raw_parser.add_argument("filepath", help="Path to the raw time-course Excel file.")
    raw_parser.add_argument(
        "--sheet-name",
        default=None,
        help="Optional Excel sheet name."
    )
    raw_parser.add_argument(
        "--calibration-factor",
        type=float,
        default=0.0056,
        help="OD/s to uM/s conversion denominator. Default: 0.0056."
    )
    raw_parser.add_argument(
        "--fit-start",
        type=float,
        default=None,
        help="First time point to use for initial-rate fitting."
    )
    raw_parser.add_argument(
        "--fit-end",
        type=float,
        default=None,
        help="Last time point to use for initial-rate fitting."
    )
    raw_parser.add_argument(
        "--max-points",
        type=int,
        default=3,
        help="Use only the first N selected time points. Use 0 for all selected points."
    )
    raw_parser.add_argument(
        "--save-prefix",
        default=None,
        help="Optional filename prefix for saving plots. By default, figures are not saved."
    )

    single_parser = subparsers.add_parser(
        "single",
        help="Fit from Excel data that already contains ATP and initial-rate columns."
    )
    single_parser.add_argument("filepath", help="Path to the Excel file.")
    single_parser.add_argument("--x-col", type=int, default=1, help="Zero-based ATP column index.")
    single_parser.add_argument("--y-col", type=int, default=2, help="Zero-based rate column index.")
    single_parser.add_argument("--sheet-name", default=None, help="Optional Excel sheet name.")
    single_parser.add_argument(
        "--save-figure",
        default=None,
        help="Optional path to save the Hill fit figure. By default, figures are not saved."
    )

    grouped_parser = subparsers.add_parser(
        "grouped",
        help="Fit grouped Excel data with columns like Condition | ATP | Rate."
    )
    grouped_parser.add_argument("filepath", help="Path to the Excel file.")
    grouped_parser.add_argument(
        "--condition-col",
        type=int,
        default=0,
        help="Zero-based condition column index."
    )
    grouped_parser.add_argument("--x-col", type=int, default=1, help="Zero-based ATP column index.")
    grouped_parser.add_argument("--y-col", type=int, default=2, help="Zero-based rate column index.")
    grouped_parser.add_argument("--sheet-name", default=None, help="Optional Excel sheet name.")
    grouped_parser.add_argument(
        "--save-figure",
        default=None,
        help="Optional path to save the combined figure. By default, figures are not saved."
    )

    return parser


def main(argv=None):
    """
    Command-line entry point.
    """

    parser = build_cli_parser()
    args = parser.parse_args(argv)

    if args.command == "raw":
        return run_hill_fit_fully_automated(
            args.filepath,
            sheet_name=args.sheet_name,
            calibration_factor=args.calibration_factor,
            fit_start=args.fit_start,
            fit_end=args.fit_end,
            max_points=None if args.max_points == 0 else args.max_points,
            output_prefix=args.save_prefix
        )

    if args.command == "single":
        return run_hill_fit_from_excel(
            args.filepath,
            x_col=args.x_col,
            y_col=args.y_col,
            sheet_name=args.sheet_name,
            output_file=args.save_figure
        )

    if args.command == "grouped":
        return run_hill_fit_multiple_from_excel(
            args.filepath,
            condition_col=args.condition_col,
            x_col=args.x_col,
            y_col=args.y_col,
            sheet_name=args.sheet_name,
            output_file=args.save_figure
        )

    parser.error(f"Unknown command: {args.command}")


if __name__ == "__main__":
    main()
